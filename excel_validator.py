#!/usr/bin/python -u
# -*- coding: UTF-8 -*-

import argparse
import os.path
import sys
import time

import yaml
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.comments import Comment
from progress.bar import Bar

from validator import *
from validator import HeaderValidator


def isValid(settings, value, coordinate, errors, value2=None):
    try:
        # validator list
        classmap = {
            'NotBlank': NotBlankValidator.NotBlankValidator,
            'Type': TypeValidator.TypeValidator,
            'Length': LengthValidator.LengthValidator,
            'Regex': RegexValidator.RegexValidator,
            'Email': EmailValidator.EmailValidator,
            'Choice': ChoiceValidator.ChoiceValidator,
            'Date': DateTimeValidator.DateTimeValidator,
            'ExcelDate': ExcelDateValidator.ExcelDateValidator,
            'Country': CountryValidator.CountryValidator,
            'Conditional': ConditionalValidator.ConditionalValidator,
            'Header': HeaderValidator.HeaderValidator
        }

        violations = []
        validator_type = list(settings.keys())[0]
        data = list(settings.values())[0]
        validator = classmap[validator_type](data)

        if validator_type != 'Conditional':
            result = validator.validate(value)
        else:
            result = validator.validate(value, value2)

        if (result == False):
            violations.append(validator.getMessage())

        if len(violations) > 0:
            if len(errors) > 0:
                find_coord_list = [x for x in errors if x[0] == coordinate]
                if len(find_coord_list) > 0:
                    coord_error = [x for x in errors if x[0] == coordinate][0]
                    coord_error_idx = errors.index(coord_error)
                    coord_error = (coordinate, coord_error[1] + violations)
                    errors[coord_error_idx] = coord_error
                    return True

            errors.append((coordinate, violations))

        return True
    except Exception as e:
        raise Exception('Exception while validating cell {}, {}: {}'.format(coordinate, type(e), str(e)))


def setSettings(config):
    settings = {}
    excludes = []

    print("Get validation config " + config)
    try:
        stream = open(config, 'r',  encoding='utf8')
        config = yaml.load(stream)
    except IOError as e:
        print(e)
        exit(1)

    if 'validators' in config and 'columns' in config.get('validators'):
        settings['validators'] = config.get('validators').get('columns')
    else:
        return False

    if 'default' in config.get('validators'):
        settings['defaultValidator'] = config.get('validators').get('default')[0]
    else:
        settings['defaultValidator'] = None

    if 'excludes' in config:
        for column in config.get('excludes'):
            excludes.append(column_index_from_string(column))
        settings['excludes'] = excludes
    else:
        settings['excludes'] = []

    settings['range'] = dict(min_row=None, max_row=None,
                             min_col=None, max_col=None)

    if 'range_col' in config:
        settings['range']['min_col'] = column_index_from_string(config.get('range_col')[0])
        settings['range']['max_col'] = column_index_from_string(config.get('range_col')[1])

    if 'range_row' in config:
        settings['range']['min_row'] = int(config.get('range_row')[0])
        settings['range']['max_row'] = int(config.get('range_row')[1])

    if 'data_from_row' in config:
        settings['data_from_row'] = int(config.get('data_from_row'))
    else:
        settings['data_from_row'] = 0

    return settings


def markErrors(errors, excelFile, sheetName, tmpDir, printErrors=False):
    progressBar = Bar('Processing', max=len(errors))

    if os.path.getsize(excelFile) > 10485760:
        print("Log broken cells")
        for error in errors:
            progressBar.next()

            if printErrors:
                print("Broken Excel cell: " + error[0] + " [ " + ','.join(error[1]) + " ]")
            else:
                print("Broken Excel cell: " + error[0])

        progressBar.finish()
        return

    # open Excel file
    newFile = os.path.join(tmpDir,
                           "errors_" + time.strftime("%Y-%m-%d") + "_" + str(int(time.time())) + "_" + os.path.basename(
                               excelFile))
    fileName, fileExtension = os.path.splitext(excelFile)

    if fileExtension == '.xlsm':
        wb = load_workbook(excelFile, keep_vba=True, data_only=True)
    else:
        wb = load_workbook(excelFile, data_only=True)

    creator = wb.properties.creator
    ws = wb.get_sheet_by_name(sheetName)

    contrast_fill = PatternFill(start_color='FFEB9C',
                          end_color='FFEB9C',
                          fill_type='solid')

    for error in errors:
        progressBar.next()

        print("Broken Excel cell: " + error[0])
        cell = ws[error[0]]
        if printErrors:
            cell.comment = Comment(','.join(error[1]), 'Автопроверка')
        cell.fill = contrast_fill
        cell.font = Font(color='9C5700')

    progressBar.finish()
    # save error excel file
    wb.properties.creator = creator
    print("[[Save file: " + newFile + "]]")
    try:
        wb.save(newFile)
    except Exception as e:
        print(e)
        exit(1)

    return newFile


def validate(settings, excelFile, sheetName, tmpDir, printErrors=False):
    print("Validate Excel Sheet " + sheetName)

    errors = []
    # open Excel file
    print("Parse Excel file")
    wb = load_workbook(excelFile, keep_vba=True, data_only=True, read_only=True)
    ws = wb.get_sheet_by_name(sheetName)

    if 'range' in settings and settings['range']['max_row'] is None:
        settings['range']['max_row'] = ws.max_row
        settings['range']['min_row'] = ws.min_row
    if 'range' in settings and settings['range']['max_col'] is None:
        settings['range']['max_col'] = ws.max_column
        settings['range']['min_col'] = ws.min_column
    print(settings['range'])

    progressBar = Bar('Processing', max=settings['range']['max_row'])

    # iterate excel sheet
    for rom_counter, row in enumerate(ws.iter_rows(**settings['range']), 1):
        progressBar.next()
        # do not parse empty rows
        if isEmpty(row):
            continue
        for cell_column, cell in enumerate(row, 1):
            coordinates = get_column_letter(cell_column) + str(rom_counter)
            try:
                value = cell.value
            except ValueError:
                errors.append((coordinates, ValueError))

            # skip excludes column
            if hasattr(cell, 'column') and cell_column in settings['excludes']:
                continue

            if get_column_letter(cell_column) in settings['validators']:
                for type in settings['validators'][get_column_letter(cell_column)]:
                    name = list(type.keys())[0]
                    if name == 'Header':
                        try:
                            row_to_check = int(list(type.values())[0]['row'])
                        except KeyError:
                            row_to_check = 1
                        if rom_counter == row_to_check:
                            isValid(type, value, coordinates, errors)
                        continue
                    elif rom_counter >= settings['data_from_row']:
                        if name == 'Conditional':
                            fieldB = list(type.values())[0]['fieldB']
                            value2 = ws[fieldB + str(rom_counter)].value
                            isValid(type, value, coordinates, errors, value2)
                            continue
                        else:
                            isValid(type, value, coordinates, errors)

            elif settings['defaultValidator'] != None and rom_counter >= settings['data_from_row']:
                isValid(settings['defaultValidator'], value, coordinates, errors)

    progressBar.finish()

    print("Found %d error(s)" % len(errors))
    if (len(errors) > 0):
        return markErrors(errors, excelFile, sheetName, tmpDir, printErrors)

    return True


def isEmpty(row):
    for cell in row:
        if cell.value:
            return False

    return True


if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='Mark validation errors in Excel sheet.')
    parser.add_argument('config', metavar='config', help='Path to YAML config file')
    parser.add_argument('file', metavar='file', help='Path to excel sheet file')
    parser.add_argument('sheetName', metavar='sheetName', help='Excel Sheet Name')
    parser.add_argument('tmpDir', metavar='tmpDir', help='Temporary directory path')
    parser.add_argument('--errors', metavar='errors', help='Print errors messages in cells marked as invalid')
    args = parser.parse_args()

    settings = setSettings(args.config)

    if settings == False:
        sys.exit("Incorrect config file " + args.config)

    '''try:
        results = validate(settings, args.file, args.sheetName, args.tmpDir, args.errors)
    except Exception as e:
        sys.exit("Error occured: " + str(e))'''
    results = validate(settings, args.file, args.sheetName, args.tmpDir, args.errors)

    if results != True:
        if results:
            sys.exit("Validation errors store in: [[" + results + "]]")
        else:
            sys.exit("Invalid file is too big to generate annotated Excel file")

    sys.exit(0)
